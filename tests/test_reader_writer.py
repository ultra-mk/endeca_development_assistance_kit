import unittest
import openpyxl
from edak import reader_writer


class EXCEL_READER_TEST(unittest.TestCase):

    @classmethod
    def setUpClass(EXCEL_READER_TEST):
        EXCEL_READER_TEST.reader = reader_writer.Excel_Reader(
            'endeca_attributes.xlsx')

    def test_init(self):
        self.assertEqual("<class 'edak.reader_writer.Excel_Reader'>", str(
            type(EXCEL_READER_TEST.reader)))


class EXCEL_WRITER_TEST(unittest.TestCase):

    @classmethod
    def setUpClass(EXCEL_WRITER_TEST):
        EXCEL_WRITER_TEST.writer = reader_writer.Excel_Writer(
            'excel_writer_test.xlsx',[['SHIP_BY_DATE', 'mdex:dateTime'], ['ORDER_ID', 'mdex:string']])
        EXCEL_WRITER_TEST.test_file = openpyxl.load_workbook(
            'excel_writer_test.xlsx')

        EXCEL_WRITER_TEST.ws = EXCEL_WRITER_TEST.test_file.active

    def test_init(self):
        self.assertEqual("<class 'edak.reader_writer.Excel_Writer'>", str(
            type(EXCEL_WRITER_TEST.writer)))

    def test_worksheet_name(self):
        self.assertEqual('endeca_attributes', EXCEL_WRITER_TEST.ws.title)

    def test_header_values(self):
        self.assertEqual(['eid_instance_id', 'eid_instance_attribute'], [
                         EXCEL_WRITER_TEST.ws['A1'].value, EXCEL_WRITER_TEST.ws['B1'].value])

    def test_column_data(self):
    	self.assertEqual('SHIP_BY_DATE', EXCEL_WRITER_TEST.ws['B2'].value)

    def test_datatype_data(self):
    	self.assertEqual('mdex:string', EXCEL_WRITER_TEST.ws['C3'].value)


class TEXT_WRITER_TEST(unittest.TestCase):

    @classmethod
    def setUpClass(TEXT_WRITER_TEST):
        TEXT_WRITER_TEST.writer = reader_writer.Text_Writer(
            'test_attribute_sql.txt')

    def test_init(self):
        self.assertEqual("<class 'edak.reader_writer.Text_Writer'>", str(
            type(TEXT_WRITER_TEST.writer)))


if __name__ == '__main__':
    unittest.main()
