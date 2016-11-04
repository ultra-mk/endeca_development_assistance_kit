import openpyxl


class Excel_Reader(object):

    def __init__(self, file_name):
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.get_sheet_by_name('endeca_attributes')
        highest_row = str(sheet.get_highest_row())
        self.attribute_data = [[c.value for c in rowOfCellObjects]
                               for rowOfCellObjects in sheet['A2':'E' + highest_row]]
        self.eql_attributes = [i[1] for i in self.attribute_data]
        self.xml_attributes = [[i[1], i[2]] for i in self.attribute_data]


class Excel_Writer(object):

    def __init__(self, file_name, columns_and_datatype):
        self.file_name = file_name
        self.columns_and_datatype = columns_and_datatype
        self.create_excel_file()

    def create_excel_file(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'endeca_attributes'
        headers = ('eid_instance_id', 'eid_instance_attribute', 'datatype', 'profile_id', 'display_name')
        ws['A1'].value, ws['B1'].value, ws['C1'].value, ws['D1'].value, ws['E1'].value = headers
        for i, e in enumerate(self.columns_and_datatype):
            ws.cell(row=i + 2, column=2).value = e[0]
            ws.cell(row=i + 2, column=3).value = e[1]
        wb.save(self.file_name)


class Text_Writer(object):

    def __init__(self, file):
        self.file = file

    def clear_file(self):
        open(self.file, 'w').close()

    def save_text(self, text):
        with open(self.file, 'a') as f:
            f.write(text)
